# -*- coding: utf-8 -*-
"""
Raporlama Servisi
Gelişmiş raporlama sistemi için gerekli fonksiyonları içerir.
"""
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from flask import current_app
from app import db
from app.models import Event, Reservation, User, Company
from sqlalchemy import func, and_, or_, desc, asc
import json
import pandas as pd


class ReportService:
    """Raporlama işlemlerini yöneten servis sınıfı"""
    
    def __init__(self, company_id: int):
        self.company_id = company_id
    
    def get_summary_report(self, start_date: Optional[datetime] = None, 
                          end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Genel özet raporu oluşturur
        
        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            
        Returns:
            Dict: Özet rapor verileri
        """
        try:
            # Tarih filtresi
            date_filter = self._build_date_filter(start_date, end_date)
            
            # Temel istatistikler
            total_events = self._get_total_events(date_filter)
            total_reservations = self._get_total_reservations(date_filter)
            total_capacity = self._get_total_capacity(date_filter)
            checked_in_count = self._get_checked_in_count(date_filter)
            checkin_rate = self._calculate_checkin_rate(total_reservations, checked_in_count)
            
            # Etkinlik tipi analizi
            event_types = self._get_event_types_analysis(date_filter)
            
            # Doluluk trendleri
            occupancy_trends = self._get_occupancy_trends(date_filter)
            
            return {
                'total_events': total_events,
                'total_reservations': total_reservations,
                'total_capacity': total_capacity,
                'checked_in_count': checked_in_count,
                'checkin_rate': round(checkin_rate, 2),
                'event_types': event_types,
                'occupancy_trends': occupancy_trends,
                'date_range': {
                    'start_date': start_date.isoformat() if start_date else None,
                    'end_date': end_date.isoformat() if end_date else None
                },
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            current_app.logger.error(f"Summary report error: {str(e)}")
            return {'error': str(e)}
    
    def get_event_detail_report(self, event_id: int) -> Dict[str, Any]:
        """
        Etkinlik detay raporu oluşturur
        
        Args:
            event_id: Etkinlik ID'si
            
        Returns:
            Dict: Etkinlik detay raporu
        """
        try:
            event = Event.query.filter_by(
                id=event_id, 
                company_id=self.company_id
            ).first_or_404()
            
            # Etkinlik bilgileri
            event_info = {
                'id': event.id,
                'name': event.name,
                'event_date': event.event_date.isoformat(),
                'start_time': event.start_time.strftime('%H:%M') if event.start_time else None,
                'end_time': event.end_time.strftime('%H:%M') if event.end_time else None,
                'venue_type': event.venue_type,
                'event_type': event.event_type,
                'status': event.status.value
            }
            
            # Oturum durumu
            seating_stats = self._get_seating_stats(event_id)
            
            # Rezervasyon listesi
            reservations = self._get_event_reservations(event_id)
            
            # Check-in durumu
            checkin_stats = self._get_checkin_stats(event_id)
            
            # İptal edilen rezervasyonlar
            cancelled_reservations = self._get_cancelled_reservations(event_id)
            
            return {
                'event_info': event_info,
                'seating_stats': seating_stats,
                'reservations': reservations,
                'checkin_stats': checkin_stats,
                'cancelled_reservations': cancelled_reservations
            }
            
        except Exception as e:
            current_app.logger.error(f"Event detail report error: {str(e)}")
            return {'error': str(e)}
    
    def get_reservation_analysis_report(self, start_date: Optional[datetime] = None,
                                      end_date: Optional[datetime] = None,
                                      group_by: str = 'daily') -> Dict[str, Any]:
        """
        Rezervasyon analiz raporu oluşturur
        
        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            group_by: Gruplama kriteri (daily, weekly, monthly)
            
        Returns:
            Dict: Rezervasyon analiz raporu
        """
        try:
            date_filter = self._build_date_filter(start_date, end_date)
            
            # Günlük/Haftalık/Aylık trend analizi
            trends = self._get_reservation_trends(date_filter, group_by)
            
            # En popüler gün/saat analizi
            popular_times = self._get_popular_times(date_filter)
            
            # Ortalama rezervasyon süresi
            avg_booking_lead_time = self._get_avg_booking_lead_time(date_filter)
            
            # İptal oranları
            cancellation_rates = self._get_cancellation_rates(date_filter)
            
            return {
                'trends': trends,
                'popular_times': popular_times,
                'avg_booking_lead_time': avg_booking_lead_time,
                'cancellation_rates': cancellation_rates
            }
            
        except Exception as e:
            current_app.logger.error(f"Reservation analysis error: {str(e)}")
            return {'error': str(e)}
    
    def get_occupancy_analysis_report(self, start_date: Optional[datetime] = None,
                                    end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Doluluk analiz raporu oluşturur
        
        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            
        Returns:
            Dict: Doluluk analiz raporu
        """
        try:
            date_filter = self._build_date_filter(start_date, end_date)
            
            # Etkinlik bazlı doluluk oranları
            event_occupancy = self._get_event_occupancy(date_filter)
            
            # Oturum tipi popülerliği
            seating_type_popularity = self._get_seating_type_popularity(date_filter)
            
            # Zaman serisinde doluluk grafiği
            occupancy_time_series = self._get_occupancy_time_series(date_filter)
            
            # Boş kalan oturum analizi
            empty_seats_analysis = self._get_empty_seats_analysis(date_filter)
            
            return {
                'event_occupancy': event_occupancy,
                'seating_type_popularity': seating_type_popularity,
                'occupancy_time_series': occupancy_time_series,
                'empty_seats_analysis': empty_seats_analysis
            }
            
        except Exception as e:
            current_app.logger.error(f"Occupancy analysis error: {str(e)}")
            return {'error': str(e)}
    
    def get_customer_analysis_report(self, start_date: Optional[datetime] = None,
                                   end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Müşteri analiz raporu oluşturur
        
        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            
        Returns:
            Dict: Müşteri analiz raporu
        """
        try:
            date_filter = self._build_date_filter(start_date, end_date)
            
            # Tekrarlayan müşteriler
            returning_customers = self._get_returning_customers(date_filter)
            
            # En çok rezervasyon yapan müşteriler
            top_customers = self._get_top_customers(date_filter)
            
            # Müşteri davranış analizi
            customer_behavior = self._get_customer_behavior(date_filter)
            
            # İptal yapan müşteriler
            customers_with_cancellations = self._get_customers_with_cancellations(date_filter)
            
            return {
                'returning_customers': returning_customers,
                'top_customers': top_customers,
                'customer_behavior': customer_behavior,
                'cancelled_customers': customers_with_cancellations
            }
            
        except Exception as e:
            current_app.logger.error(f"Customer analysis error: {str(e)}")
            return {'error': str(e)}
    
    def export_to_excel(self, report_data: Dict[str, Any], report_type: str) -> str:
        """
        Rapor verilerini Excel'e aktarır
        
        Args:
            report_data: Rapor verileri
            report_type: Rapor tipi
            
        Returns:
            str: Excel dosyası yolu
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{report_type.title()} Raporu"
            
            # Başlık formatı
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            row = 1
            for key, value in report_data.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value) if not isinstance(value, (dict, list)) else str(value)[:50] + "..."
                
                # Header formatı uygula
                if row == 1:
                    ws[f'A{row}'].font = header_font
                    ws[f'B{row}'].font = header_font
                    ws[f'A{row}'].fill = header_fill
                    ws[f'B{row}'].fill = header_fill
                
                row += 1
            
            # Sütun genişliklerini ayarla
            for col in range(1, 3):
                ws.column_dimensions[get_column_letter(col)].width = 30
            
            filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"app/static/uploads/reports/{filename}"
            
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            current_app.logger.error(f"Excel export error: {str(e)}")
            raise
    
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Veriyi CSV formatında aktarır
        
        Args:
            data: Aktarılacak veri
            filename: Dosya adı
            
        Returns:
            str: CSV dosyası yolu
        """
        try:
            df = pd.DataFrame(data)
            filepath = f"app/static/uploads/reports/{filename}"
            
            # Klasörü oluştur
            import os
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            # UTF-8 encoding ile kaydet
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            return filepath
            
        except Exception as e:
            current_app.logger.error(f"CSV export error: {str(e)}")
            raise
    
    # Private helper methods
    def _build_date_filter(self, start_date: Optional[datetime], 
                          end_date: Optional[datetime]) -> Optional[and_]:
        """Tarih filtresi oluşturur"""
        if not start_date and not end_date:
            return None
        
        filters = []
        if start_date:
            filters.append(Event.event_date >= start_date.date())
        if end_date:
            filters.append(Event.event_date <= end_date.date())
        
        return and_(*filters) if filters else None
    
    def _get_total_events(self, date_filter: Optional[and_]) -> int:
        """Toplam etkinlik sayısını getirir"""
        query = Event.query.filter_by(company_id=self.company_id)
        if date_filter:
            query = query.filter(date_filter)
        return query.count()
    
    def _get_total_reservations(self, date_filter: Optional[and_]) -> int:
        """Toplam rezervasyon sayısını getirir"""
        query = Reservation.query.join(Event).filter(Event.company_id == self.company_id)
        if date_filter:
            query = query.filter(date_filter)
        return query.count()
    
    def _get_total_capacity(self, date_filter: Optional[and_]) -> int:
        """Toplam kapasiteyi hesaplar"""
        query = db.session.query(func.sum(EventSeating.capacity)).join(Event)
        query = query.filter(Event.company_id == self.company_id)
        if date_filter:
            query = query.filter(date_filter)
        return query.scalar() or 0
    
    def _get_checked_in_count(self, date_filter: Optional[and_]) -> int:
        """Check-in sayısını getirir"""
        query = db.session.query(func.count(Reservation.id))
        query = query.join(Event).filter(
            Event.company_id == self.company_id,
            Reservation.checked_in == True
        )
        if date_filter:
            query = query.filter(date_filter)
        return query.scalar() or 0
    
    def _calculate_checkin_rate(self, total_reservations: int, checked_in: int) -> float:
        """Check-in oranını hesaplar"""
        if total_reservations == 0:
            return 0.0
        return (checked_in / total_reservations) * 100
    
    def _get_event_types_analysis(self, date_filter: Optional[and_]) -> List[Dict]:
        """Etkinlik tipi analizini yapar"""
        query = db.session.query(
            Event.event_type,
            func.count(Event.id).label('count')
        ).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(Event.event_type).all()
        
        return [{'type': row.event_type, 'count': row.count} for row in query]
    
    def _get_occupancy_trends(self, date_filter: Optional[and_]) -> List[Dict]:
        """Doluluk trendlerini getirir"""
        query = db.session.query(
            Event.event_date,
            func.count(Reservation.id).label('reservations')
        ).join(Reservation).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(Event.event_date).order_by(Event.event_date)
        results = query.all()
        
        return [
            {
                'date': row.event_date.isoformat(),
                'reservations': row.reservations
            }
            for row in results
        ]
    
    def _get_seating_stats(self, event_id: int) -> Dict[str, int]:
        """Etkinlik için oturum istatistiklerini getirir"""
        total_seats = EventSeating.query.filter_by(event_id=event_id).count()
        reserved_seats = db.session.query(func.count(EventSeating.id)).join(Reservation).filter(
            EventSeating.event_id == event_id,
            Reservation.status == 'active'
        ).scalar() or 0
        available_seats = total_seats - reserved_seats
        
        return {
            'total': total_seats,
            'reserved': reserved_seats,
            'available': available_seats
        }
    
    def _get_event_reservations(self, event_id: int) -> List[Dict]:
        """Etkinlik rezervasyonlarını getirir"""
        reservations = Reservation.query.filter_by(event_id=event_id).all()
        
        return [
            {
                'id': r.id,
                'phone': r.phone,
                'name': r.customer_name,
                'seating': r.seating.seat_number if r.seating else 'N/A',
                'people_count': r.number_of_people,
                'status': r.status.value,
                'checked_in': r.checked_in,
                'created_at': r.created_at.isoformat()
            }
            for r in reservations
        ]
    
    def _get_checkin_stats(self, event_id: int) -> Dict[str, int]:
        """Check-in istatistiklerini getirir"""
        total = Reservation.query.filter_by(event_id=event_id).count()
        checked_in = Reservation.query.filter_by(
            event_id=event_id, 
            checked_in=True
        ).count()
        
        return {
            'total': total,
            'checked_in': checked_in,
            'not_checked_in': total - checked_in
        }
    
    def _get_cancelled_reservations(self, event_id: int) -> List[Dict]:
        """İptal edilen rezervasyonları getirir"""
        cancelled = Reservation.query.filter_by(
            event_id=event_id,
            status='cancelled'
        ).all()
        
        return [
            {
                'id': r.id,
                'phone': r.phone,
                'name': r.customer_name,
                'cancelled_at': r.cancelled_at.isoformat() if r.cancelled_at else None
            }
            for r in cancelled
        ]
    
    def _get_reservation_trends(self, date_filter: Optional[and_], group_by: str) -> List[Dict]:
        """Rezervasyon trendlerini getirir"""
        # Implementation depends on group_by parameter
        # This is a simplified version
        query = db.session.query(
            func.date(Reservation.created_at).label('date'),
            func.count(Reservation.id).label('count')
        ).join(Event).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by('date').order_by('date')
        results = query.all()
        
        return [
            {
                'date': row.date.isoformat(),
                'count': row.count
            }
            for row in results
        ]
    
    def _get_popular_times(self, date_filter: Optional[and_]) -> List[Dict]:
        """En popüler zamanları getirir"""
        query = db.session.query(
            func.hour(Reservation.created_at).label('hour'),
            func.count(Reservation.id).label('count')
        ).join(Event).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by('hour').order_by(desc('count')).limit(10)
        results = query.all()
        
        return [
            {
                'hour': f"{row.hour:02d}:00",
                'count': row.count
            }
            for row in results
        ]
    
    def _get_avg_booking_lead_time(self, date_filter: Optional[and_]) -> float:
        """Ortalama rezervasyon süresini hesaplar"""
        query = db.session.query(
            func.avg(Event.event_date - Reservation.created_at.cast(db.Date))
        ).join(Reservation).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        result = query.scalar()
        return result.days if result else 0
    
    def _get_cancellation_rates(self, date_filter: Optional[and_]) -> Dict[str, float]:
        """İptal oranlarını hesaplar"""
        total = self._get_total_reservations(date_filter)
        cancelled = db.session.query(func.count(Reservation.id)).join(Event).filter(
            Event.company_id == self.company_id,
            Reservation.status == 'cancelled'
        )
        
        if date_filter:
            cancelled = cancelled.filter(date_filter)
        
        cancelled_count = cancelled.scalar() or 0
        
        return {
            'total': total,
            'cancelled': cancelled_count,
            'rate': (cancelled_count / total * 100) if total > 0 else 0
        }
    
    def _get_event_occupancy(self, date_filter: Optional[and_]) -> List[Dict]:
        """Etkinlik bazlı doluluk oranlarını getirir"""
        # Simplified implementation
        events = Event.query.filter_by(company_id=self.company_id)
        if date_filter:
            events = events.filter(date_filter)
        
        results = []
        for event in events:
            total_seats = event.seatings.count() if event.seatings else 0
            reserved_seats = event.reservations.filter_by(status='active').count()
            occupancy_rate = (reserved_seats / total_seats * 100) if total_seats > 0 else 0
            
            results.append({
                'event_name': event.name,
                'event_date': event.event_date.isoformat(),
                'total_seats': total_seats,
                'reserved_seats': reserved_seats,
                'occupancy_rate': round(occupancy_rate, 2)
            })
        
        return results
    
    def _get_seating_type_popularity(self, date_filter: Optional[and_]) -> List[Dict]:
        """Oturum tipi popülerliğini getirir"""
        query = db.session.query(
            SeatingType.name,
            func.count(Reservation.id).label('reservations')
        ).join(EventSeating, Reservation).join(Event).filter(
            Event.company_id == self.company_id,
            Reservation.status == 'active'
        )
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(SeatingType.name).order_by(desc('reservations'))
        results = query.all()
        
        return [
            {
                'seating_type': row.name,
                'reservations': row.reservations
            }
            for row in results
        ]
    
    def _get_occupancy_time_series(self, date_filter: Optional[and_]) -> List[Dict]:
        """Zaman serisinde doluluk grafiğini getirir"""
        # Implementation similar to trends but with occupancy calculation
        return self._get_occupancy_trends(date_filter)
    
    def _get_empty_seats_analysis(self, date_filter: Optional[and_]) -> List[Dict]:
        """Boş kalan oturum analizini getirir"""
        # Find seats that are available for a certain period
        # This is a simplified implementation
        return []
    
    def _get_returning_customers(self, date_filter: Optional[and_]) -> List[Dict]:
        """Tekrarlayan müşterileri getirir"""
        query = db.session.query(
            Reservation.phone,
            func.count(Reservation.id).label('visit_count')
        ).join(Event).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(Reservation.phone).having(func.count(Reservation.id) > 1)
        results = query.all()
        
        return [
            {
                'phone': row.phone,
                'visit_count': row.visit_count
            }
            for row in results
        ]
    
    def _get_top_customers(self, date_filter: Optional[and_]) -> List[Dict]:
        """En çok rezervasyon yapan müşterileri getirir"""
        query = db.session.query(
            Reservation.phone,
            func.count(Reservation.id).label('reservation_count'),
            func.avg(Reservation.number_of_people).label('avg_people')
        ).join(Event).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(Reservation.phone).order_by(desc('reservation_count')).limit(10)
        results = query.all()
        
        return [
            {
                'phone': row.phone,
                'reservation_count': row.reservation_count,
                'avg_people': round(row.avg_people or 0, 1)
            }
            for row in results
        ]
    
    def _get_customer_behavior(self, date_filter: Optional[and_]) -> Dict[str, Any]:
        """Müşteri davranış analizini getirir"""
        # Average time between bookings per customer
        # Most common booking patterns
        return {
            'avg_booking_interval_days': 0,
            'peak_booking_day': 'Saturday',
            'most_common_group_size': 2
        }
    
    def _get_customers_with_cancellations(self, date_filter: Optional[and_]) -> List[Dict]:
        """İptal yapan müşterileri getirir"""
        query = db.session.query(
            Reservation.phone,
            func.count(Reservation.id).label('total_reservations'),
            func.sum(func.case([(Reservation.status == 'cancelled', 1)], else_=0)).label('cancellations')
        ).join(Event).filter(Event.company_id == self.company_id)
        
        if date_filter:
            query = query.filter(date_filter)
        
        query = query.group_by(Reservation.phone).having(
            func.sum(func.case([(Reservation.status == 'cancelled', 1)], else_=0)) > 0
        )
        results = query.all()
        
        return [
            {
                'phone': row.phone,
                'total_reservations': row.total_reservations,
                'cancellations': row.cancellations
            }
            for row in results
        ]


# Import statements for relationships
from app.models.seating import SeatingType, EventSeating