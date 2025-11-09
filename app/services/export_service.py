"""
Export Service - PDF, Excel, CSV dışa aktarma işlemleri
Gelişmiş raporlama sistemi için export fonksiyonları
"""

from datetime import datetime
from io import BytesIO
import csv
import json
from flask import send_file, make_response
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from .analytics_service import AnalyticsService


class ExportService:
    """Verileri çeşitli formatlarda dışa aktaran servis"""
    
    def __init__(self):
        self.analytics_service = AnalyticsService()
    
    def export_to_csv(self, event_id, file_format='utf-8'):
        """CSV formatında dışa aktar"""
        try:
            df = self.analytics_service.export_data_to_dataframe(event_id)
            
            # CSV oluştur
            output = BytesIO()
            df.to_csv(output, index=False, encoding=file_format)
            output.seek(0)
            
            # Dosya adı
            event_name = self._get_event_name(event_id)
            filename = f"rezervasyon_raporu_{event_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            raise Exception(f"CSV export hatası: {str(e)}")
    
    def export_to_excel(self, event_id, include_analytics=True):
        """Excel formatında dışa aktar"""
        try:
            workbook = Workbook()
            
            # Ana sayfa - Rezervasyonlar
            ws_reservations = workbook.active
            ws_reservations.title = "Rezervasyonlar"
            
            # Verileri al
            df = self.analytics_service.export_data_to_dataframe(event_id)
            
            # Başlıklar ve veriler
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_reservations.append(r)
            
            # Stil uygula
            self._style_excel_sheet(ws_reservations, "Rezervasyonlar")
            
            # Analitik sayfaları
            if include_analytics:
                self._add_analytics_sheets(workbook, event_id)
            
            # Excel dosyasını oluştur
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Dosya adı
            event_name = self._get_event_name(event_id)
            filename = f"detayli_rapor_{event_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            raise Exception(f"Excel export hatası: {str(e)}")
    
    def export_to_pdf(self, event_id, report_type='comprehensive'):
        """PDF formatında dışa aktar"""
        try:
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Başlık stili
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Etkinlik bilgileri
            event = self._get_event(event_id)
            
            # PDF Başlığı
            title = Paragraph(f"Etkinlik Raporu: {event.name}", title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Tarih
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1
            )
            date_text = Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style)
            story.append(date_text)
            story.append(Spacer(1, 20))
            
            if report_type == 'comprehensive':
                # Kapsamlı rapor
                self._add_pdf_overview(story, event_id, styles)
                self._add_pdf_trends(story, event_id, styles)
                self._add_pdf_seating_analysis(story, event_id, styles)
                self._add_pdf_customer_analysis(story, event_id, styles)
            
            elif report_type == 'summary':
                # Özet rapor
                self._add_pdf_summary(story, event_id, styles)
            
            elif report_type == 'reservations':
                # Sadece rezervasyonlar
                self._add_pdf_reservations(story, event_id, styles)
            
            # PDF oluştur
            doc.build(story)
            output.seek(0)
            
            # Dosya adı
            event_name = self._get_event_name(event_id)
            filename = f"pdf_rapor_{event_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            
            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            raise Exception(f"PDF export hatası: {str(e)}")
    
    def _style_excel_sheet(self, worksheet, sheet_name):
        """Excel sayfasına stil uygula"""
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık satırı
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Veri satırları
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
        
        # Sütun genişlikleri
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_analytics_sheets(self, workbook, event_id):
        """Excel'e analitik sayfaları ekle"""
        # Oturum analizi
        ws_seating = workbook.create_sheet("Oturum Analizi")
        seating_data = self.analytics_service.get_seating_analysis(event_id)
        
        # Başlıklar
        headers = ['Oturum No', 'Oturum Tip', 'Kapasite', 'Rezervasyon', 'Doluluk %', 'Durum']
        ws_seating.append(headers)
        
        # Veriler
        for seating in seating_data:
            ws_seating.append([
                seating['seat_number'],
                seating['seating_type'],
                seating['capacity'],
                seating['reservations'],
                f"{seating['occupancy_rate']:.1f}%",
                seating['status']
            ])
        
        self._style_excel_sheet(ws_seating, "Oturum Analizi")
        
        # Müşteri analizi
        ws_customer = workbook.create_sheet("Müşteri Analizi")
        customer_data = self.analytics_service.get_customer_analysis(event_id)
        
        # Müşteri segmentasyonu
        segments_headers = ['Segment', 'Adet']
        ws_customer.append(["Müşteri Segmentasyonu", ""])
        ws_customer.append(segments_headers)
        
        for segment, count in customer_data['segments'].items():
            ws_customer.append([segment.replace('_', ' ').title(), count])
        
        ws_customer.append([])
        ws_customer.append(["Müşteri Detayları", ""])
        ws_customer.append(headers)
        
        for customer in customer_data['customer_list']:
            ws_customer.append([
                customer['phone'],
                customer['name'],
                customer['reservation_count'],
                customer['total_people']
            ])
        
        self._style_excel_sheet(ws_customer, "Müşteri Analizi")
    
    def _add_pdf_overview(self, story, event_id, styles):
        """PDF'e genel bakış ekle"""
        analytics = self.analytics_service.get_event_overview_analytics(event_id)
        
        # Başlık
        story.append(Paragraph("Genel Bakış", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Veri tablosu
        data = [
            ['Metrik', 'Değer'],
            ['Toplam Kapasite', str(analytics['capacity_metrics']['total_capacity'])],
            ['Aktif Rezervasyon', str(analytics['capacity_metrics']['active_reservations'])],
            ['Boş Koltuk', str(analytics['capacity_metrics']['available_seats'])],
            ['Doluluk Oranı', f"{analytics['capacity_metrics']['occupancy_rate']}%"],
            ['Check-in Oranı', f"{analytics['reservation_metrics']['checkin_rate']}%"]
        ]
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    def _add_pdf_trends(self, story, event_id, styles):
        """PDF'e trend analizi ekle"""
        trends = self.analytics_service.get_reservation_trends(event_id, days=7)
        
        story.append(Paragraph("Rezervasyon Trendleri (Son 7 Gün)", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if trends['daily_trends']:
            data = [['Tarih', 'Rezervasyon', 'Check-in']]
            for trend in trends['daily_trends']:
                data.append([
                    trend['date'],
                    str(trend['reservations']),
                    str(trend['checkins'])
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("Veri bulunamadı", styles['Normal']))
        
        story.append(Spacer(1, 20))
    
    def _add_pdf_seating_analysis(self, story, event_id, styles):
        """PDF'e oturum analizi ekle"""
        seating = self.analytics_service.get_seating_analysis(event_id)
        
        story.append(Paragraph("Oturum Doluluk Analizi", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if seating:
            data = [['Oturum No', 'Tip', 'Kapasite', 'Rezervasyon', 'Doluluk %']]
            for seat in seating[:10]:  # İlk 10 oturum
                data.append([
                    seat['seat_number'],
                    seat['seating_type'],
                    str(seat['capacity']),
                    str(seat['reservations']),
                    f"{seat['occupancy_rate']:.1f}%"
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("Oturum verisi bulunamadı", styles['Normal']))
        
        story.append(Spacer(1, 20))
    
    def _add_pdf_customer_analysis(self, story, event_id, styles):
        """PDF'e müşteri analizi ekle"""
        customers = self.analytics_service.get_customer_analysis(event_id)
        
        story.append(Paragraph("Müşteri Analizi", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Segmentasyon
        story.append(Paragraph("Müşteri Segmentasyonu:", styles['Heading3']))
        segments_text = f"""
        Yeni Müşteriler: {customers['segments']['new_customers']}<br/>
        Tekrar Eden Müşteriler: {customers['segments']['returning_customers']}<br/>
        Grup Rezervasyonları: {customers['segments']['group_bookings']}<br/>
        Bireysel Rezervasyonlar: {customers['segments']['individual_bookings']}<br/>
        Toplam Benzersiz Müşteri: {customers['total_unique_customers']}
        """
        story.append(Paragraph(segments_text, styles['Normal']))
        story.append(Spacer(1, 20))
    
    def _add_pdf_summary(self, story, event_id, styles):
        """PDF'e özet ekle"""
        # Kısa özet rapor
        analytics = self.analytics_service.get_event_overview_analytics(event_id)
        
        story.append(Paragraph("Özet Rapor", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        summary_text = f"""
        <b>Etkinlik:</b> {analytics['event_info']['name']}<br/>
        <b>Toplam Kapasite:</b> {analytics['capacity_metrics']['total_capacity']}<br/>
        <b>Aktif Rezervasyon:</b> {analytics['capacity_metrics']['active_reservations']}<br/>
        <b>Doluluk Oranı:</b> {analytics['capacity_metrics']['occupancy_rate']}%<br/>
        <b>Check-in Oranı:</b> {analytics['reservation_metrics']['checkin_rate']}%
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
    
    def _add_pdf_reservations(self, story, event_id, styles):
        """PDF'e rezervasyon listesi ekle"""
        df = self.analytics_service.export_data_to_dataframe(event_id)
        
        story.append(Paragraph("Rezervasyon Listesi", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        if not df.empty:
            # İlk 20 rezervasyon
            limited_df = df.head(20)
            
            data = [['Ad Soyad', 'Telefon', 'Kişi Sayısı', 'Durum', 'Check-in']]
            for _, row in limited_df.iterrows():
                data.append([
                    f"{row['Ad']} {row['Soyad']}",
                    row['Telefon'],
                    str(row['Kisi_Sayisi']),
                    row['Durum'],
                    row['Check_in']
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        story.append(Paragraph(f"Toplam {len(df)} rezervasyon", styles['Normal']))
    
    def _get_event(self, event_id):
        """Etkinlik bilgisini al"""
        from app.models import Event
        return Event.query.get_or_404(event_id)
    
    def _get_event_name(self, event_id):
        """Etkinlik adını al"""
        event = self._get_event(event_id)
        # Dosya adı için güvenli karakterler
        safe_name = "".join(c for c in event.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        return safe_name.replace(' ', '_')
    
    def create_json_export(self, event_id):
        """JSON formatında tam veri export"""
        try:
            # Tüm analitik verileri topla
            overview = self.analytics_service.get_event_overview_analytics(event_id)
            trends = self.analytics_service.get_reservation_trends(event_id)
            seating = self.analytics_service.get_seating_analysis(event_id)
            customers = self.analytics_service.get_customer_analysis(event_id)
            timing = self.analytics_service.get_time_based_analysis(event_id)
            
            # JSON yapısı
            export_data = {
                'export_info': {
                    'event_id': event_id,
                    'export_date': datetime.now().isoformat(),
                    'export_type': 'comprehensive'
                },
                'event_overview': overview,
                'reservation_trends': trends,
                'seating_analysis': seating,
                'customer_analysis': customers,
                'timing_analysis': timing
            }
            
            # JSON dosyası oluştur
            output = BytesIO()
            json_data = json.dumps(export_data, ensure_ascii=False, indent=2, default=str)
            output.write(json_data.encode('utf-8'))
            output.seek(0)
            
            # Dosya adı
            event_name = self._get_event_name(event_id)
            filename = f"tam_veri_{event_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            
            return send_file(
                output,
                mimetype='application/json',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            raise Exception(f"JSON export hatası: {str(e)}")
